"""
Container Stowage Optimizer - GENETIC ALGORITHM ONLY  v5-GA
============================================================
Versi ini HANYA menggunakan Genetic Algorithm (GA) tanpa Simulated Annealing.
Tujuan: membandingkan kualitas solusi GA murni vs SA murni vs GA+SA gabungan.

Parameter identik dengan v5-RANDOM (GA+SA):
  - pop_size=80, generations=50
  - Crossover + 4 jenis mutasi
  - Inisialisasi: random_valid_solution()

TIDAK ADA SA inner-loop pada setiap child.
"""

import numpy as np
import random
import time
import sys
import os
from typing import List, Tuple, Set, Dict, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================================
# SHIP & SLOT CONFIG
# ============================================================================
LIGHTSHIP = {'weight': 2943.63, 'lcg': 41.49, 'vcg': 8.65, 'tcg': 0.0}
TANKS = [
    {'name': 'Tank', 'weight': 0.0, 'lcg': 0.5, 'vcg': 1.0, 'tcg': 0.0},
]

BAY_LCGS  = {1:76.929, 3:72.459, 5:66.349, 7:59.699, 9:53.589,
             11:46.469, 13:40.349, 15:33.719, 17:27.589, 19:20.969, 21:13.479}
ROW_TCGS  = {0:0.0, 1:1.22, 2:-1.22, 3:3.66, 4:-3.66, 5:6.1, 6:-6.1}
TIER_VCGS = {2:1.527, 4:4.118, 6:6.709,
             82:9.300, 84:11.340, 86:13.931, 88:16.522, 90:19.113, 92:21.704}

BAYS  = sorted(BAY_LCGS.keys())
ROWS  = sorted(ROW_TCGS.keys())
TIERS = sorted(TIER_VCGS.keys())

NB, NR, NT = len(BAYS), len(ROWS), len(TIERS)

def bi2n(i): return BAYS[i]
def ri2n(i): return ROWS[i]
def ti2n(i): return TIERS[i]

VCG_TARGET = 9.0
LCG_TARGET = 2.0

TOL_VCG = 0.5
TOL_TCG = 0.1
TOL_LCG = 1.0

ALLOWED_40FT_PAIRS: Set[int] = {1, 3, 5, 7}
PAIR_SECOND_BAYS: Set[int] = {p + 1 for p in ALLOWED_40FT_PAIRS if p + 1 < NB}

# ============================================================================
# LOCK_20FT_ONLY_SLOTS
# ============================================================================
_LOCKED_BAY3_BAY5_R3R6: Set[Tuple[int,int,int]] = {
    (1,3,1),(1,3,2),(1,3,4),(1,3,5),(1,3,6),
    (1,4,1),(1,4,4),(1,4,5),(1,4,6),
    (1,5,2),(1,5,4),(1,5,5),(1,5,6),
    (1,6,2),(1,6,4),(1,6,5),(1,6,6),
    (2,3,0),(2,3,1),(2,3,2),(2,3,4),(2,3,5),(2,3,6),(2,3,7),
    (2,4,0),(2,4,1),(2,4,2),(2,4,4),(2,4,5),(2,4,6),(2,4,7),
    (2,5,1),(2,5,2),(2,5,4),(2,5,5),(2,5,6),(2,5,7),
    (2,6,1),(2,6,2),(2,6,4),(2,6,5),(2,6,6),(2,6,7),
}

_LOCKED_ASYMMETRIC: Set[Tuple[int,int,int]] = {
    (2,0,7),(2,1,7),(2,2,7),
    (4,6,4),(4,6,5),(4,6,6),(4,6,7),
    (8,6,4),(8,6,5),(8,6,6),(8,6,7),
}

LOCK_20FT_ONLY_SLOTS: Set[Tuple[int,int,int]] = (
    _LOCKED_BAY3_BAY5_R3R6 | _LOCKED_ASYMMETRIC
)

_LOCKED_ROWS_FOR_PAIR: Dict[int, Set[int]] = {
    1: {3, 4, 5, 6},
    2: {3, 4, 5, 6},
}

def is_locked_20ft_only(b: int, r: int, t: int) -> bool:
    return (b, r, t) in LOCK_20FT_ONLY_SLOTS

def is_row_locked_for_pair(b: int, r: int) -> bool:
    return r in _LOCKED_ROWS_FOR_PAIR.get(b, set())

# ============================================================================
# VALID SLOTS
# ============================================================================
VALID_SLOTS: Set[Tuple[int,int,int]] = {
    (0,0,4),(0,0,5),(0,0,6),(0,0,7),(0,1,4),(0,1,5),(0,1,6),(0,1,7),(0,2,4),(0,2,5),
    (0,2,6),(0,2,7),(0,3,4),(0,3,5),(0,3,6),(0,3,7),(0,4,4),(0,4,5),(0,4,6),(0,4,7),
    (1,0,4),(1,0,5),(1,0,6),(1,1,0),(1,1,1),(1,1,2),(1,1,4),(1,1,5),(1,1,6),(1,2,0),
    (1,2,1),(1,2,2),(1,2,4),(1,2,5),(1,2,6),(1,3,1),(1,3,2),(1,3,4),(1,3,5),(1,3,6),
    (1,4,1),(1,4,4),(1,4,5),(1,4,6),(1,5,2),(1,5,4),(1,5,5),(1,5,6),(1,6,2),(1,6,4),
    (1,6,5),(1,6,6),
    (2,0,4),(2,0,5),(2,0,6),(2,0,7),(2,1,0),(2,1,1),(2,1,2),(2,1,4),(2,1,5),(2,1,6),
    (2,1,7),(2,2,0),(2,2,1),(2,2,2),(2,2,4),(2,2,5),(2,2,6),(2,2,7),(2,3,0),(2,3,1),
    (2,3,2),(2,3,4),(2,3,5),(2,3,6),(2,3,7),(2,4,0),(2,4,1),(2,4,2),(2,4,4),(2,4,5),
    (2,4,6),(2,4,7),(2,5,1),(2,5,2),(2,5,4),(2,5,5),(2,5,6),(2,5,7),(2,6,1),(2,6,2),
    (2,6,4),(2,6,5),(2,6,6),(2,6,7),
    (3,0,4),(3,0,5),(3,0,6),(3,0,7),(3,1,0),(3,1,1),(3,1,2),(3,1,4),(3,1,5),(3,1,6),
    (3,1,7),(3,2,0),(3,2,1),(3,2,2),(3,2,4),(3,2,5),(3,2,6),(3,2,7),(3,3,0),(3,3,1),
    (3,3,2),(3,3,4),(3,3,5),(3,3,6),(3,3,7),(3,4,0),(3,4,1),(3,4,2),(3,4,4),(3,4,5),
    (3,4,6),(3,4,7),(3,5,0),(3,5,1),(3,5,2),(3,5,4),(3,5,5),(3,5,6),(3,5,7),(3,6,0),
    (3,6,1),(3,6,2),
    (4,0,4),(4,0,5),(4,0,6),(4,0,7),(4,1,0),(4,1,1),(4,1,2),(4,1,4),(4,1,5),(4,1,6),
    (4,1,7),(4,2,0),(4,2,1),(4,2,2),(4,2,4),(4,2,5),(4,2,6),(4,2,7),(4,3,0),(4,3,1),
    (4,3,2),(4,3,4),(4,3,5),(4,3,6),(4,3,7),(4,4,0),(4,4,1),(4,4,2),(4,4,4),(4,4,5),
    (4,4,6),(4,4,7),(4,5,0),(4,5,1),(4,5,2),(4,5,4),(4,5,5),(4,5,6),(4,5,7),(4,6,0),
    (4,6,1),(4,6,2),(4,6,4),(4,6,5),(4,6,6),(4,6,7),
    (5,0,4),(5,0,5),(5,0,6),(5,0,7),(5,1,0),(5,1,1),(5,1,2),(5,1,4),(5,1,5),(5,1,6),
    (5,1,7),(5,2,0),(5,2,1),(5,2,2),(5,2,4),(5,2,5),(5,2,6),(5,2,7),(5,3,0),(5,3,1),
    (5,3,2),(5,3,4),(5,3,5),(5,3,6),(5,3,7),(5,4,0),(5,4,1),(5,4,2),(5,4,4),(5,4,5),
    (5,4,6),(5,4,7),(5,5,0),(5,5,1),(5,5,2),(5,5,4),(5,5,5),(5,5,6),(5,5,7),(5,6,0),
    (5,6,1),(5,6,2),(5,6,4),(5,6,5),(5,6,6),(5,6,7),
    (6,0,4),(6,0,5),(6,0,6),(6,0,7),(6,1,0),(6,1,1),(6,1,2),(6,1,4),(6,1,5),(6,1,6),
    (6,1,7),(6,2,0),(6,2,1),(6,2,2),(6,2,4),(6,2,5),(6,2,6),(6,2,7),(6,3,0),(6,3,1),
    (6,3,2),(6,3,4),(6,3,5),(6,3,6),(6,3,7),(6,4,0),(6,4,1),(6,4,2),(6,4,4),(6,4,5),
    (6,4,6),(6,4,7),(6,5,0),(6,5,1),(6,5,2),(6,5,4),(6,5,5),(6,5,6),(6,5,7),(6,6,0),
    (6,6,1),(6,6,2),(6,6,4),(6,6,5),(6,6,6),(6,6,7),
    (7,0,4),(7,0,5),(7,0,6),(7,0,7),(7,1,0),(7,1,1),(7,1,2),(7,1,4),(7,1,5),(7,1,6),
    (7,1,7),(7,2,0),(7,2,1),(7,2,2),(7,2,4),(7,2,5),(7,2,6),(7,2,7),(7,3,0),(7,3,1),
    (7,3,2),(7,3,4),(7,3,5),(7,3,6),(7,3,7),(7,4,0),(7,4,1),(7,4,2),(7,4,4),(7,4,5),
    (7,4,6),(7,4,7),(7,5,0),(7,5,1),(7,5,2),(7,5,4),(7,5,5),(7,5,6),(7,5,7),(7,6,0),
    (7,6,1),(7,6,2),
    (8,0,4),(8,0,5),(8,0,6),(8,0,7),(8,1,0),(8,1,1),(8,1,2),(8,1,4),(8,1,5),(8,1,6),
    (8,1,7),(8,2,0),(8,2,1),(8,2,2),(8,2,4),(8,2,5),(8,2,6),(8,2,7),(8,3,0),(8,3,1),
    (8,3,2),(8,3,4),(8,3,5),(8,3,6),(8,3,7),(8,4,0),(8,4,1),(8,4,2),(8,4,4),(8,4,5),
    (8,4,6),(8,4,7),(8,5,0),(8,5,1),(8,5,2),(8,5,4),(8,5,5),(8,5,6),(8,5,7),(8,6,0),
    (8,6,1),(8,6,2),(8,6,4),(8,6,5),(8,6,6),(8,6,7),
    (9,0,4),(9,0,5),(9,0,6),(9,0,7),(9,0,8),(9,1,0),(9,1,1),(9,1,2),(9,1,4),(9,1,5),
    (9,1,6),(9,1,7),(9,1,8),(9,2,0),(9,2,1),(9,2,2),(9,2,4),(9,2,5),(9,2,6),(9,2,7),
    (9,2,8),(9,3,0),(9,3,1),(9,3,2),(9,3,4),(9,3,5),(9,3,6),(9,3,7),(9,3,8),(9,4,0),
    (9,4,1),(9,4,2),(9,4,4),(9,4,5),(9,4,6),(9,4,7),(9,4,8),(9,5,1),(9,5,2),(9,5,4),
    (9,5,5),(9,5,6),(9,5,7),(9,5,8),(9,6,1),(9,6,2),(9,6,4),(9,6,5),(9,6,6),
    (10,0,3),(10,0,4),(10,0,5),(10,0,6),(10,0,7),(10,0,8),(10,1,3),(10,1,4),(10,1,5),(10,1,6),
    (10,1,7),(10,1,8),(10,2,3),(10,2,4),(10,2,5),(10,2,6),(10,2,7),(10,2,8),(10,3,3),(10,3,4),
    (10,3,5),(10,3,6),(10,3,7),(10,3,8),(10,4,2),(10,4,3),(10,4,4),(10,4,5),(10,4,6),(10,4,7),
    (10,4,8),
}

TOTAL_SLOTS = len(VALID_SLOTS)

COL_TIERS: Dict[Tuple[int,int], List[int]] = {}
for b in range(NB):
    for r in range(NR):
        ts = sorted(t for t in range(NT) if (b,r,t) in VALID_SLOTS)
        if ts:
            COL_TIERS[(b,r)] = ts

PAIR_TIERS: Dict[Tuple[int,int], List[int]] = {}
for pidx in ALLOWED_40FT_PAIRS:
    if pidx+1 >= NB: continue
    for r in range(NR):
        if is_row_locked_for_pair(pidx, r):
            continue
        ts = sorted(t for t in range(NT)
                    if (pidx,r,t) in VALID_SLOTS
                    and (pidx+1,r,t) in VALID_SLOTS
                    and not is_locked_20ft_only(pidx, r, t)
                    and not is_locked_20ft_only(pidx+1, r, t))
        if ts:
            PAIR_TIERS[(pidx, r)] = ts

SLOTS_BY_TIER = [[(b,r) for (b,r,t) in VALID_SLOTS if t == ti] for ti in range(NT)]
PAIR_BY_TIER  = [[(b,r) for b in ALLOWED_40FT_PAIRS for r in range(NR)
                  if (b,r,t) in VALID_SLOTS
                  and (b+1,r,t) in VALID_SLOTS
                  and not is_row_locked_for_pair(b, r)
                  and not is_locked_20ft_only(b, r, t)
                  and not is_locked_20ft_only(b+1, r, t)]
                 for t in range(NT)]

_n_locked_total = len(LOCK_20FT_ONLY_SLOTS)

print(f"""
==============================================================
  CONTAINER STOWAGE - GA ONLY  v5
  Hanya Genetic Algorithm (tanpa Simulated Annealing)
  Untuk perbandingan: GA vs SA vs GA+SA
--------------------------------------------------------------
  Valid slots : {TOTAL_SLOTS:<6}  Target VCG={VCG_TARGET}m  LCG={LCG_TARGET}m
  40ft pairs  : Bay {', '.join(f'{bi2n(p)}-{bi2n(p+1)}' for p in sorted(ALLOWED_40FT_PAIRS)[:4])}
  Locked 20ft : {_n_locked_total:<3} slots
==============================================================
""")

# ============================================================================
# CONTAINER
# ============================================================================
class Container:
    __slots__ = ('id','weight','size','name')
    def __init__(self, cid, weight, size=1, name=None):
        self.id = cid
        self.weight = weight
        self.size = size
        if name:
            self.name = name
        else:
            prefix = 'CB' if size == 2 else 'CA'
            self.name = f"{prefix}{cid:03d}"

    def __repr__(self):
        return f"{self.name}({self.weight:.0f}t)"

# ============================================================================
# READ INPUT
# ============================================================================
def read_containers(filename: str) -> List[Container]:
    df = pd.read_excel(filename)
    df.columns = df.columns.str.lower().str.strip().str.replace(' ','_')

    wcol = next((c for c in ['berat','weight','berat(ton)','weight(ton)'] if c in df.columns), None)
    scol = next((c for c in ['size','ukuran','teu','ft'] if c in df.columns), None)
    ncol = next((c for c in ['nomor_kontainer','container','container_id','nama','name'] if c in df.columns), None)

    if not wcol:
        raise ValueError("Kolom berat/weight tidak ditemukan")

    containers = []
    for idx, row in df.iterrows():
        try:
            weight = float(row[wcol])
            size = 1
            name = None
            if scol and pd.notna(row[scol]):
                sv = str(row[scol]).lower()
                size = 2 if ('40' in sv or sv.strip() == '2') else 1
            if ncol and pd.notna(row[ncol]):
                raw = str(row[ncol]).strip()
                name = raw
                if name[:2].upper() == 'CB':
                    size = 2
            containers.append(Container(len(containers)+1, weight, size, name))
        except Exception as e:
            print(f"  Warn row {idx+2}: {e}")

    c20 = sum(1 for c in containers if c.size == 1)
    c40 = sum(1 for c in containers if c.size == 2)
    ratio = c40 / (c20 + c40) * 100 if (c20 + c40) > 0 else 0
    print(f"  {len(containers)} containers | 20ft:{c20} | 40ft:{c40} | "
          f"Rasio 40ft: {ratio:.1f}%")
    return containers

# ============================================================================
# STOWAGE ARRAY
# ============================================================================
def empty_stowage() -> np.ndarray:
    return np.zeros((NB, NR, NT), dtype=np.int16)

def col_type(s, cmap, b, r):
    sizes = {cmap[s[b,r,t]].size for t in range(NT) if s[b,r,t] > 0}
    if not sizes: return 'empty'
    if len(sizes) > 1:
        top_20_ti = -1; bot_40_ti = NT
        for ti in range(NT):
            cid = s[b, r, ti]
            if cid > 0:
                if cmap[cid].size == 1: top_20_ti = max(top_20_ti, ti)
                else: bot_40_ti = min(bot_40_ti, ti)
        return 'pair_split' if top_20_ti < bot_40_ti else 'mixed'
    return '20ft' if 1 in sizes else '40ft'

def col_type_simple(s, cmap, b, r):
    return col_type(s, cmap, b, r)

def count_40under20(s, cmap):
    violations = 0; seen40 = set()
    for b in range(NB):
        for r in range(NR):
            ts = COL_TIERS.get((b,r), [])
            if not ts: continue
            top_20 = -1
            for ti in reversed(ts):
                cid = s[b,r,ti]
                if cid > 0 and cmap[cid].size == 1: top_20 = ti; break
            if top_20 == -1: continue
            for ti in ts:
                if ti >= top_20: break
                cid = s[b,r,ti]
                if cid > 0 and cmap[cid].size == 2 and cid not in seen40:
                    seen40.add(cid); violations += 1
    return violations

def count_lock_violations(s, cmap):
    violations = 0; seen40 = set()
    for (b,r,t) in LOCK_20FT_ONLY_SLOTS:
        cid = s[b,r,t]
        if cid > 0 and cmap[cid].size == 2 and cid not in seen40:
            seen40.add(cid); violations += 1
    return violations

def compute_balance_penalty(s, cmap):
    total_20 = sum(1 for c in cmap.values() if c.size == 1)
    total_40 = sum(1 for c in cmap.values() if c.size == 2)
    if total_20 + total_40 == 0: return 0.0
    target_ratio_40 = total_40 / (total_20 + total_40)
    mid = NB // 2; zones = [(0, mid), (mid, NB)]; penalty = 0.0; c40_seen = set()
    for z_start, z_end in zones:
        z_20 = z_40 = 0
        for b in range(z_start, z_end):
            for r in range(NR):
                for t in range(NT):
                    cid = s[b,r,t]
                    if cid <= 0: continue
                    c = cmap[cid]
                    if c.size == 1: z_20 += 1
                    elif cid not in c40_seen: c40_seen.add(cid); z_40 += 1
        z_total = z_20 + z_40
        if z_total == 0: continue
        deviation = abs(z_40/z_total - target_ratio_40)
        penalty += deviation * z_total
    return penalty

def rebuild(col20, pair40, cmap, strict_40over20=True):
    s = empty_stowage(); unplaced = []
    for pidx in ALLOWED_40FT_PAIRS:
        if pidx+1 >= NB: continue
        for r in range(NR):
            if is_row_locked_for_pair(pidx, r): continue
            ts_pair = [t for t in PAIR_TIERS.get((pidx,r), [])
                       if not is_locked_20ft_only(pidx,r,t)
                       and not is_locked_20ft_only(pidx+1,r,t)]
            if not ts_pair: continue
            cids_40 = sorted(pair40.get((pidx,r), []), key=lambda c: -cmap[c].weight)
            for i, cid in enumerate(cids_40[:len(ts_pair)]):
                t = ts_pair[i]
                s[pidx,r,t] = cid; s[pidx+1,r,t] = cid
            if len(cids_40) > len(ts_pair):
                unplaced.extend(cids_40[len(ts_pair):])

    pair_free: Dict[Tuple[int,int], List[int]] = {}
    for pidx in ALLOWED_40FT_PAIRS:
        if pidx+1 >= NB: continue
        for r in range(NR):
            ts_pair = PAIR_TIERS.get((pidx,r), [])
            if is_row_locked_for_pair(pidx,r):
                pair_free[(pidx,r)] = [ti for ti in COL_TIERS.get((pidx,r),[]) if s[pidx,r,ti]==0]
                pair_free[(pidx+1,r)] = [ti for ti in COL_TIERS.get((pidx+1,r),[]) if s[pidx+1,r,ti]==0]
                continue
            if not ts_pair: continue
            low40 = NT
            for ti in range(NT):
                if s[pidx,r,ti] > 0 and cmap[s[pidx,r,ti]].size == 2: low40 = ti; break
            ff = [ti for ti in range(NT) if ti < low40 and (pidx,r,ti) in VALID_SLOTS and s[pidx,r,ti]==0]
            if ff: pair_free[(pidx,r)] = ff
            fs = [ti for ti in COL_TIERS.get((pidx+1,r),[]) if ti < low40 and s[pidx+1,r,ti]==0]
            if fs: pair_free[(pidx+1,r)] = fs

    all_20 = sorted({cid for cids in col20.values() for cid in cids}, key=lambda cid: -cmap[cid].weight)
    non_pair = [(b,r,ti) for (b,r),ts in COL_TIERS.items()
                if b not in ALLOWED_40FT_PAIRS and b not in PAIR_SECOND_BAYS
                for ti in ts if s[b,r,ti]==0]
    pair_slots = [(b,r,ti) for (b,r),tis in pair_free.items() for ti in tis]
    avail = non_pair + pair_slots; placed_set = set()

    for cid in all_20:
        placed = False
        for slot in avail:
            b,r,ti = slot
            if slot in placed_set or s[b,r,ti] != 0: continue
            ts_col = COL_TIERS.get((b,r),[])
            ok = all(not (s[b,r,tc] > 0 and cmap.get(s[b,r,tc], Container(0,0)).size == 2)
                     for tc in ts_col if tc < ti)
            if not ok: continue
            placed_set.add(slot); s[b,r,ti] = cid; placed = True; break
        if not placed: unplaced.append(cid)
    return s, unplaced

def stowage_to_cols(s, cmap):
    c40_ids = {c.id for c in cmap.values() if c.size == 2}
    pair40: Dict = {}; seen40 = set()
    for pidx in ALLOWED_40FT_PAIRS:
        if pidx+1 >= NB: continue
        for r in range(NR):
            if is_row_locked_for_pair(pidx, r): continue
            ts = PAIR_TIERS.get((pidx,r), []); cids = []
            for t in ts:
                cid = s[pidx,r,t]
                if cid > 0 and cid in c40_ids and cid not in seen40:
                    seen40.add(cid); cids.append(cid)
            if cids: pair40[(pidx,r)] = cids
    col20: Dict = {}
    for b in range(NB):
        for r in range(NR):
            ts = COL_TIERS.get((b,r), [])
            cids = [s[b,r,t] for t in ts if s[b,r,t] > 0 and s[b,r,t] not in c40_ids]
            if cids: col20[(b,r)] = cids
    return col20, pair40

def fix_floating(s, cmap):
    col20, pair40 = stowage_to_cols(s, cmap)
    result, _ = rebuild(col20, pair40, cmap, strict_40over20=True)
    return result

# ============================================================================
# RANDOM VALID SOLUTION
# ============================================================================
def random_valid_solution(containers):
    cmap = {c.id: c for c in containers}
    c40_list = [c for c in containers if c.size == 2]
    c20_list = [c for c in containers if c.size == 1]
    random.shuffle(c40_list); random.shuffle(c20_list)

    pair_keys = list(PAIR_TIERS.keys()); random.shuffle(pair_keys)
    pair40: Dict = {}; c40_remaining = list(c40_list)

    for key in pair_keys:
        if not c40_remaining: break
        pidx, r = key; cap = len(PAIR_TIERS[key])
        max_here = min(cap, len(c40_remaining))
        n_here   = random.randint(0, max_here)
        if n_here > 0:
            assigned = [c40_remaining.pop(0).id for _ in range(n_here)]
            pair40[key] = assigned

    if c40_remaining:
        for key in pair_keys:
            if not c40_remaining: break
            pidx, r = key; cap = len(PAIR_TIERS[key]); existing = len(pair40.get(key, []))
            space = cap - existing
            if space > 0:
                extra = [c40_remaining.pop(0).id for _ in range(min(space, len(c40_remaining)))]
                pair40.setdefault(key, []).extend(extra)

    all_col_keys = list(COL_TIERS.keys()); random.shuffle(all_col_keys)
    col20: Dict = {}; c20_remaining = list(c20_list)

    for key in all_col_keys:
        if not c20_remaining: break
        b, r = key; cap = len(COL_TIERS[key])
        if b in ALLOWED_40FT_PAIRS and not is_row_locked_for_pair(b, r):
            cap = max(0, cap - len(pair40.get((b, r), [])))
        elif b in PAIR_SECOND_BAYS:
            cap = max(0, cap - len(pair40.get((b-1, r), [])))
        if cap <= 0: continue
        max_here = min(cap, len(c20_remaining))
        n_here   = random.randint(1, max_here) if max_here > 0 else 0
        if n_here > 0:
            assigned = [c20_remaining.pop(0).id for _ in range(n_here)]
            col20[key] = assigned

    if c20_remaining:
        for key in all_col_keys:
            if not c20_remaining: break
            b, r = key; cap = len(COL_TIERS[key]); existing = len(col20.get(key, []))
            if b in ALLOWED_40FT_PAIRS and not is_row_locked_for_pair(b, r):
                cap = max(0, cap - len(pair40.get((b, r), [])))
            elif b in PAIR_SECOND_BAYS:
                cap = max(0, cap - len(pair40.get((b-1, r), [])))
            space = cap - existing
            if space > 0:
                extra = [c20_remaining.pop(0).id for _ in range(min(space, len(c20_remaining)))]
                col20.setdefault(key, []).extend(extra)

    s, unplaced = rebuild(col20, pair40, cmap, strict_40over20=True)
    return s, cmap

# ============================================================================
# FITNESS
# ============================================================================
_BASE_OFFSET = 2000

def _stab_penalty_teu(value, target, a_lin, a_quad, tier2_teu, tier3_teu,
                      thresh2, thresh3, tol=0.0):
    d = max(0.0, abs(value - target) - tol)
    p = a_lin * d + a_quad * d * d
    if d > thresh2: p += tier2_teu
    if d > thresh3: p += tier3_teu
    return p

def fitness(s, cmap):
    LPP = 93.6; midship = LPP / 2.0
    lw = LIGHTSHIP['weight']; tw = sum(t['weight'] for t in TANKS)
    base_w  = lw + tw
    base_mv = lw * LIGHTSHIP['vcg'] + sum(t['weight'] * t['vcg'] for t in TANKS)
    base_ml = lw * (midship - LIGHTSHIP['lcg']) + sum(t['weight'] * (midship - t['lcg']) for t in TANKS)
    cw = cmv = cml = cmt = 0.0; proc = set(); n20 = n40 = 0
    for b in range(NB):
        for r in range(NR):
            for t in range(NT):
                cid = s[b, r, t]
                if cid <= 0 or cid in proc: continue
                proc.add(cid); c = cmap[cid]
                vcg = TIER_VCGS[ti2n(t)]; tcg = ROW_TCGS[ri2n(r)]; bn = bi2n(b)
                if c.size == 1:
                    lcg_ap = BAY_LCGS[bn]; n20 += 1
                else:
                    bn2 = bi2n(b+1) if b+1 < NB else bn
                    lcg_ap = (BAY_LCGS[bn] + BAY_LCGS[bn2]) / 2; n40 += 1
                w = c.weight; cw += w; cmv += w * vcg; cml += w * (midship - lcg_ap); cmt += w * tcg
    tot_w = base_w + cw
    if tot_w == 0: return float('-inf'), 0., 0., 0.
    vcg_val = (base_mv + cmv) / tot_w
    lcg_val = (base_ml + cml) / tot_w
    tcg_val = cmt / tot_w
    total_teu = n20 + (n40 * 2); reward = float(total_teu)
    v_penalty = _stab_penalty_teu(vcg_val, 9.0, 80.0, 60.0, 200.0, 400.0, 0.5, 1.0)
    l_penalty = _stab_penalty_teu(lcg_val, 2.0, 40.0, 20.0, 60.0, 150.0, 0.3, 1.0)
    t_penalty = _stab_penalty_teu(tcg_val, 0.0, 20.0, 15.0, 30.0, 80.0, 0.3, 1.0)
    unified_fit = _BASE_OFFSET + reward - (v_penalty + l_penalty + t_penalty)
    return unified_fit, vcg_val, lcg_val, tcg_val

# ============================================================================
# MUTATION & CROSSOVER
# ============================================================================
def get_occupied(s):
    occ = {}
    for b in range(NB):
        for r in range(NR):
            for t in range(NT):
                cid = s[b,r,t]
                if cid > 0 and cid not in occ: occ[cid] = (b,r,t)
    return occ

def mutate_swap(s, cmap):
    occ = get_occupied(s)
    ids20 = [k for k in occ if cmap[k].size == 1]
    ids40 = [k for k in occ if cmap[k].size == 2]
    pool  = ids20 if (random.random() < 0.65 and len(ids20) >= 2) else ids40
    if len(pool) < 2: return s
    id1, id2 = random.sample(pool, 2); p1, p2 = occ[id1], occ[id2]
    if cmap[id1].size == 2 and is_locked_20ft_only(*p2): return s
    if cmap[id2].size == 2 and is_locked_20ft_only(*p1): return s
    ns = s.copy(); ns[p1[0],p1[1],p1[2]] = id2; ns[p2[0],p2[1],p2[2]] = id1
    if cmap[id1].size == 2:
        ns[p1[0]+1,p1[1],p1[2]] = id2; ns[p2[0]+1,p2[1],p2[2]] = id1
    return fix_floating(ns, cmap)

def mutate_move(s, cmap):
    occ = get_occupied(s)
    ids20 = [k for k in occ if cmap[k].size == 1]
    if not ids20: return s
    cid = random.choice(ids20); b0,r0,t0 = occ[cid]
    empty_slots = [(b,r,t) for (b,r,t) in VALID_SLOTS
                   if s[b,r,t] == 0 and col_type_simple(s,cmap,b,r) in ('empty','20ft','pair_split')]
    if not empty_slots: return s
    b1,r1,t1 = random.choice(empty_slots)
    ns = s.copy(); ns[b0,r0,t0] = 0; ns[b1,r1,t1] = cid
    return fix_floating(ns, cmap)

def mutate_reorder_col(s, cmap):
    ns = s.copy(); b = random.randint(0,NB-1); r = random.randint(0,NR-1)
    ts = COL_TIERS.get((b,r), [])
    if len(ts) < 2: return ns
    c40_ids = {c.id for c in cmap.values() if c.size == 2}
    cids20  = [ns[b,r,t] for t in ts if ns[b,r,t] > 0 and ns[b,r,t] not in c40_ids]
    if len(cids20) < 2: return ns
    random.shuffle(cids20); fi = 0
    for t in ts:
        if ns[b,r,t] in c40_ids: continue
        if fi < len(cids20): ns[b,r,t] = cids20[fi]; fi += 1
        else: ns[b,r,t] = 0
    return fix_floating(ns, cmap)

def mutate_balance(s, cmap):
    occ   = get_occupied(s)
    ids20 = [k for k in occ if cmap[k].size == 1]
    if len(ids20) < 2: return s
    mid = NB // 2
    def zr(bays):
        z20 = z40 = 0; s40 = set()
        for b in bays:
            for r in range(NR):
                for t in range(NT):
                    cid = s[b,r,t]
                    if cid <= 0: continue
                    if cmap[cid].size == 1: z20 += 1
                    elif cid not in s40: s40.add(cid); z40 += 1
        return z20, z40
    f20,f40 = zr(range(mid)); a20,a40 = zr(range(mid,NB))
    fr = f20/(f20+f40) if (f20+f40)>0 else 0
    ar = a20/(a20+a40) if (a20+a40)>0 else 0
    if abs(fr-ar) < 0.05: return mutate_swap(s, cmap)
    if fr > ar:
        src = [k for k in ids20 if occ[k][0] < mid]
        tgt = [(b,r,t) for (b,r,t) in VALID_SLOTS if b >= mid and s[b,r,t] == 0
               and col_type_simple(s,cmap,b,r) in ('empty','20ft','pair_split')]
    else:
        src = [k for k in ids20 if occ[k][0] >= mid]
        tgt = [(b,r,t) for (b,r,t) in VALID_SLOTS if b < mid and s[b,r,t] == 0
               and col_type_simple(s,cmap,b,r) in ('empty','20ft','pair_split')]
    if not src or not tgt: return mutate_swap(s, cmap)
    cid = random.choice(src); b0,r0,t0 = occ[cid]; b1,r1,t1 = random.choice(tgt)
    ns  = s.copy(); ns[b0,r0,t0] = 0; ns[b1,r1,t1] = cid
    return fix_floating(ns, cmap)

def crossover(pa, pb, cmap):
    cut   = random.randint(1, NB-2)
    child = np.zeros_like(pa); child[:cut] = pa[:cut]; child[cut:] = pb[cut:]
    seen  = {}; dupes = []
    for b in range(NB):
        for r in range(NR):
            for t in range(NT):
                cid = child[b,r,t]
                if cid <= 0: continue
                if cid in seen: child[b,r,t] = 0; dupes.append(cid)
                else: seen[cid] = (b,r,t)
    missing = (set(cmap.keys()) - set(seen.keys())) | set(dupes)
    for cid in missing:
        c = cmap[cid]; placed = False
        if c.size == 2:
            for ti in range(NT):
                for (b,r) in PAIR_BY_TIER[ti]:
                    if is_locked_20ft_only(b,r,ti) or is_locked_20ft_only(b+1,r,ti): continue
                    if child[b,r,ti] == 0 and child[b+1,r,ti] == 0:
                        ct = col_type(child, cmap, b, r)
                        if ct not in ('empty','40ft'): continue
                        child[b,r,ti] = cid; child[b+1,r,ti] = cid; placed = True; break
                if placed: break
        else:
            for ti in range(NT):
                for (b,r) in SLOTS_BY_TIER[ti]:
                    if child[b,r,ti] == 0:
                        ct = col_type(child, cmap, b, r)
                        if ct not in ('empty','20ft','pair_split'): continue
                        child[b,r,ti] = cid; placed = True; break
                if placed: break
    return fix_floating(child, cmap)

# ============================================================================
# GA-ONLY OPTIMIZER  (tanpa SA inner-loop)
# ============================================================================
def optimize_ga(containers, pop_size=80, generations=50, verbose=True):
    """
    Pure Genetic Algorithm - tanpa SA inner-loop.
    Parameter identik dengan v5 asli (pop=80, gen=50),
    hanya sa_steps dihilangkan.
    """
    cmap  = {c.id: c for c in containers}
    n40   = sum(1 for c in containers if c.size == 2)
    n20   = sum(1 for c in containers if c.size == 1)

    print(f"\n{'='*65}")
    print(f"  GA ONLY | pop={pop_size}  gen={generations}")
    print(f"  {n20}x20ft + {n40}x40ft  |  Locked: {_n_locked_total} slots")
    print(f"  [SA inner-loop DINONAKTIFKAN - hanya GA murni]")
    print(f"{'='*65}")

    # Inisialisasi populasi
    print(f"\n  Membuat {pop_size} solusi awal (random valid)...")
    population = []
    for i in range(pop_size):
        s, _ = random_valid_solution(containers)
        if i > 0:
            for _ in range(random.randint(3, 10)):
                r = random.random()
                if r < 0.4:   s = mutate_swap(s, cmap)
                elif r < 0.7: s = mutate_move(s, cmap)
                else:         s = mutate_balance(s, cmap)
        population.append(s)

    scores = [fitness(s, cmap) for s in population]
    fits   = [sc[0] for sc in scores]
    print(f"\n  Fitness populasi awal:")
    print(f"    Min  = {min(fits):>10.1f}")
    print(f"    Max  = {max(fits):>10.1f}")
    print(f"    Mean = {sum(fits)/len(fits):>10.1f}")

    best_idx = max(range(pop_size), key=lambda i: scores[i][0])
    best_s   = population[best_idx].copy()
    best_fit = scores[best_idx][0]
    best_vcg = scores[best_idx][1]
    best_lcg = scores[best_idx][2]

    history  = []   # rekam fitness terbaik tiap generasi

    print(f"\n  Mulai evolusi GA (tanpa SA)...")
    for gen in range(generations):
        new_pop = []; new_scores = []
        ranked  = sorted(range(pop_size), key=lambda i: scores[i][0], reverse=True)

        # Elitism: 2 individu terbaik langsung lolos
        elite = [population[ranked[0]].copy(), population[ranked[1]].copy()]
        new_pop   += elite
        new_scores += [scores[ranked[0]], scores[ranked[1]]]

        while len(new_pop) < pop_size:
            # Tournament selection
            t1, t2 = random.sample(range(pop_size), 2)
            pa = population[t1 if scores[t1][0] > scores[t2][0] else t2]
            t3, t4 = random.sample(range(pop_size), 2)
            pb = population[t3 if scores[t3][0] > scores[t4][0] else t4]

            # Crossover
            child = crossover(pa, pb, cmap) if random.random() < 0.7 else pa.copy()

            # Mutasi (tanpa SA - langsung hitung fitness)
            mr = random.random()
            if mr < 0.30:   child = mutate_swap(child, cmap)
            elif mr < 0.55: child = mutate_move(child, cmap)
            elif mr < 0.75: child = mutate_reorder_col(child, cmap)
            else:           child = mutate_balance(child, cmap)

            child = fix_floating(child, cmap)
            cs    = fitness(child, cmap)
            new_pop.append(child); new_scores.append(cs)

        population = new_pop; scores = new_scores
        cb = max(range(pop_size), key=lambda i: scores[i][0])
        if scores[cb][0] > best_fit:
            best_fit = scores[cb][0]; best_s = population[cb].copy()
            best_vcg = scores[cb][1]; best_lcg = scores[cb][2]

        history.append(best_fit)

        if verbose and (gen % 10 == 0 or gen == generations - 1):
            v40u = count_40under20(best_s, cmap); vl = count_lock_violations(best_s, cmap)
            bal  = compute_balance_penalty(best_s, cmap)
            occ  = sum(1 for (b,r,t) in VALID_SLOTS if best_s[b,r,t] != 0)
            print(f"  Gen{gen:4d} | Fit={best_fit:>9.1f} | Fill={occ}/{TOTAL_SLOTS}({occ/TOTAL_SLOTS*100:.0f}%) | "
                  f"VCG={best_vcg:.3f}(D{abs(best_vcg-VCG_TARGET):.2f}) "
                  f"LCG={best_lcg:.3f}(D{abs(best_lcg-LCG_TARGET):.2f}) | "
                  f"40<20={v40u} Lock={vl} Bal={bal:.2f}")

    print(f"\n{'='*65}\n  SELESAI GA - Best fitness = {best_fit:.1f}\n{'='*65}")

    # Ringkasan konvergensi
    if history:
        improv_gens = sum(1 for i in range(1, len(history)) if history[i] > history[i-1])
        print(f"\n  Konvergensi: {improv_gens}/{generations} generasi menghasilkan perbaikan")
        print(f"  Fitness awal -> akhir : {history[0]:.1f} -> {history[-1]:.1f}  "
              f"(D{history[-1]-history[0]:+.1f})")

    return best_s, best_fit, history

# ============================================================================
# VALIDATION, REPORT, EXPORT  (sama dengan versi asli)
# ============================================================================
def validate(s, cmap):
    c40_ids = {c.id for c in cmap.values() if c.size == 2}
    err_float=[]; err_weight=[]; err_mixed=[]; err_40bay=[]; err_40under=[]; err_lock=[]
    for b in range(NB):
        for r in range(NR):
            ts = COL_TIERS.get((b,r), []); ct = col_type(s, cmap, b, r)
            if ct == 'mixed': err_mixed.append(f"Bay{bi2n(b)} Row{ri2n(r)}")
            top20 = -1
            for ti in reversed(ts):
                cid = s[b,r,ti]
                if cid > 0 and cmap[cid].size == 1: top20 = ti; break
            if top20 >= 0:
                for ti in ts:
                    if ti >= top20: break
                    cid = s[b,r,ti]
                    if cid > 0 and cmap[cid].size == 2:
                        err_40under.append(f"Bay{bi2n(b)} Row{ri2n(r)} T{ti2n(ti)}")
            for ti in ts:
                cid = s[b,r,ti]
                if cid > 0 and cmap[cid].size == 2 and is_locked_20ft_only(b,r,ti):
                    err_lock.append(f"Bay{bi2n(b)} Row{ri2n(r)} T{ti2n(ti)}")
            for i in range(1, len(ts)):
                ca_ = s[b,r,ts[i]]; cb_ = s[b,r,ts[i-1]]
                if ca_ > 0 and cb_ == 0: err_float.append(f"Bay{bi2n(b)} R{ri2n(r)} T{ti2n(ts[i])}")
                if ca_ > 0 and cb_ > 0:
                    if cmap[ca_].weight > cmap[cb_].weight + 2.:
                        err_weight.append(f"Bay{bi2n(b)} R{ri2n(r)} T{ti2n(ts[i-1])}->{ti2n(ts[i])}")
    seen = set()
    for b in range(NB):
        for r in range(NR):
            for t in range(NT):
                cid = s[b,r,t]
                if cid > 0 and cid not in seen and cmap[cid].size == 2:
                    seen.add(cid)
                    if b not in ALLOWED_40FT_PAIRS: err_40bay.append(f"C{cid:03d} di Bay{bi2n(b)}")
                    elif b+1 >= NB or s[b+1,r,t] != cid: err_40bay.append(f"C{cid:03d} tidak span pair")
    total = len(err_float)+len(err_weight)+len(err_mixed)+len(err_40bay)+len(err_40under)+len(err_lock)
    print(f"\n{'='*60}\n  VALIDASI - {total} pelanggaran\n{'='*60}")
    for tag, lst in [("Floating",err_float),("Urutan Berat",err_weight),("Mixed",err_mixed),
                     ("Bay 40ft",err_40bay),("40<20",err_40under),("Locked",err_lock)]:
        st = "OK" if not lst else "FAIL"
        print(f"  [{st:^4}] {tag:<14}: {len(lst)}")
        for e in lst[:3]: print(f"         x {e}")
    print(f"{'='*60}")
    return total == 0

def cog_report(s, cmap):
    _, vcg, lcg, tcg = fitness(s, cmap)
    occ = sum(1 for (b,r,t) in VALID_SLOTS if s[b,r,t] != 0)
    _p20 = set(); _p40 = set()
    for b in range(NB):
        for r in range(NR):
            for t in range(NT):
                cid = s[b,r,t]
                if cid <= 0: continue
                if cmap[cid].size == 1: _p20.add(cid)
                else: _p40.add(cid)
    n20 = len(_p20); n40 = len(_p40)
    v40u = count_40under20(s, cmap); vl = count_lock_violations(s, cmap); bal = compute_balance_penalty(s, cmap)
    print(f"\n  -- HASIL AKHIR (GA Only) ----------------------------")
    print(f"  Slot terisi : {occ}/{TOTAL_SLOTS} ({occ/TOTAL_SLOTS*100:.1f}%)")
    print(f"  Container   : {len(cmap)} (20ft:{n20}, 40ft:{n40})")
    print(f"  VCG={vcg:.3f}m (D{abs(vcg-VCG_TARGET):.3f})  LCG={lcg:.3f}m (D{abs(lcg-LCG_TARGET):.3f})  TCG={tcg:.3f}m")
    print(f"  40ft<20ft   : {v40u}  Lock viol: {vl}  Balance: {bal:.3f}")
    return vcg, lcg, tcg, n20, n40

def get_positions(s, cmap):
    pos = []; proc = set()
    for b in range(NB):
        for r in range(NR):
            for t in range(NT):
                cid = s[b,r,t]
                if cid <= 0 or cid in proc: continue
                proc.add(cid); c = cmap[cid]; bn,rn,tn = bi2n(b),ri2n(r),ti2n(t)
                vcg = TIER_VCGS[tn]; tcg = ROW_TCGS[rn]
                if c.size == 1: lcg = BAY_LCGS[bn]; bay_str = str(bn)
                else:
                    bn2 = bi2n(b+1) if b+1 < NB else bn
                    lcg = (BAY_LCGS[bn]+BAY_LCGS[bn2])/2; bay_str = f"{bn}-{bn2}"
                pos.append({'Container_ID':c.name,'Size':f'{c.size*20}ft','Weight(ton)':c.weight,
                            'Bay':bay_str,'Row':rn,'Tier':tn,'VCG(m)':round(vcg,3),
                            'LCG(m)':round(lcg,3),'TCG(m)':round(tcg,3),
                            'MomVCG':round(c.weight*vcg,3),'MomLCG':round(c.weight*lcg,3),
                            'Notes':'locked_20ft' if is_locked_20ft_only(b,r,t) else ('pair_bay' if (b in ALLOWED_40FT_PAIRS or b in PAIR_SECOND_BAYS) else '')})
    return pos

def export_excel(s, cmap, vcg, lcg, tcg, filename):
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Stowage Plan"
    HDR = Font(bold=True, color="FFFFFF")
    BG  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    BD  = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    AL  = Alignment(horizontal='center')
    ws['A1'] = "CONTAINER STOWAGE PLAN - GA ONLY v5"
    ws['A1'].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells('A1:L1'); ws['A1'].alignment = AL
    hdrs = ['Container_ID','Size','Weight(ton)','Bay','Row','Tier','VCG(m)','LCG(m)','TCG(m)','MomVCG','MomLCG','Notes']
    for ci, h in enumerate(hdrs, 1):
        cl = ws.cell(row=3, column=ci, value=h)
        cl.font = HDR; cl.fill = BG; cl.alignment = AL; cl.border = BD
    positions = get_positions(s, cmap)
    for ri, p in enumerate(positions, 4):
        for ci, key in enumerate(hdrs, 1):
            cl = ws.cell(row=ri, column=ci, value=p.get(key,''))
            cl.border = BD
            if ci >= 7 and ci <= 11: cl.number_format = '0.000'
    sr = len(positions)+6
    ws.cell(row=sr, column=1, value="SUMMARY - GA Only").font = Font(bold=True, size=12); sr += 1
    occ  = sum(1 for (b,r,t) in VALID_SLOTS if s[b,r,t] != 0)
    v40u = count_40under20(s, cmap); vl = count_lock_violations(s, cmap); bal = compute_balance_penalty(s, cmap)
    for lbl, val, note in [
        ("VCG(m)", vcg, f"Target={VCG_TARGET} D={abs(vcg-VCG_TARGET):.3f}"),
        ("LCG(m)", lcg, f"Target={LCG_TARGET} D={abs(lcg-LCG_TARGET):.3f}"),
        ("TCG(m)", tcg, "Target=0"),
        ("Slots", occ, f"/{TOTAL_SLOTS} ({occ/TOTAL_SLOTS*100:.1f}%)"),
        ("40ft<20ft", v40u, "harus 0"), ("Lock viol", vl, "harus 0"),
        ("Balance", bal, "lebih kecil lebih baik"),
        ("Algoritma", "GA Only", "Tanpa SA inner-loop"),
    ]:
        ws.cell(row=sr, column=1, value=lbl).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=round(val,3) if isinstance(val,float) else val)
        ws.cell(row=sr, column=3, value=note)
        if lbl in ("40ft<20ft","Lock viol") and isinstance(val, (int,float)) and val > 0:
            ws.cell(row=sr, column=2).font = Font(bold=True, color="FF0000")
        sr += 1
    for ci in range(1, len(hdrs)+1):
        col_letter = get_column_letter(ci); mx = 8
        for row in ws.iter_rows(min_col=ci, max_col=ci):
            for cell in row:
                try:
                    if cell.value: mx = max(mx, len(str(cell.value)))
                except: pass
        ws.column_dimensions[col_letter].width = min(mx+2, 30)
    wb.save(filename)
    print(f"  Output: {filename}")

def print_unified_summary(label, n20, n40, vcg, tcg, lcg):
    total_teu = n20 + n40 * 2; reward = float(total_teu)
    v_p = _stab_penalty_teu(vcg, 9.0, 80.0, 60.0, 200.0, 400.0, 0.5, 1.0)
    l_p = _stab_penalty_teu(lcg, 2.0, 40.0, 20.0, 60.0, 150.0, 0.3, 1.0)
    t_p = _stab_penalty_teu(tcg, 0.0, 20.0, 15.0, 30.0, 80.0, 0.3, 1.0)
    unified = _BASE_OFFSET + reward - (v_p + l_p + t_p)
    def _ind(val, target, ok=0.3, warn=1.0):
        d = abs(val-target)
        return "[OK]" if d < ok else ("[WARN]" if d < warn else "[FAIL]")
    print(); print("+" + "="*62 + "+")
    print(f"|  {'UNIFIED FITNESS SUMMARY':^58}  |")
    print(f"|  Program : {label:<50}|")
    print("+" + "="*62 + "+")
    print(f"|  TEU dimuat  : {total_teu:>5}  ({n20}x20ft + {n40}x40ft){'':<19}|")
    print(f"|  VCG         : {vcg:>8.3f} m   target 9.000 m  D{abs(vcg-9.0):.3f}  {_ind(vcg,9.0)}  |")
    print(f"|  TCG         : {tcg:>8.4f} m   target 0.000 m  D{abs(tcg):.4f}  {_ind(tcg,0.0)}  |")
    print(f"|  LCG         : {lcg:>8.3f} m   target 2.000 m  D{abs(lcg-2.0):.3f}  {_ind(lcg,2.0)}  |")
    print("+" + "-"*62 + "+")
    print(f"|  Penalty VCG : {v_p:>8.2f} TEU                              |")
    print(f"|  Penalty LCG : {l_p:>8.2f} TEU                              |")
    print(f"|  Penalty TCG : {t_p:>8.2f} TEU                              |")
    print(f"|  Reward TEU  : {reward:>8.1f}                                 |")
    print(f"|  Offset      : {_BASE_OFFSET:>8.1f}                                 |")
    print("+" + "="*62 + "+")
    print(f"|  * UNIFIED FITNESS = {unified:>10.2f}                        |")
    print("+" + "="*62 + "+")

# ============================================================================
# MAIN
# ============================================================================
if __name__ == "__main__":
    t0 = time.time()
    random.seed(42); np.random.seed(42)

    input_file = "500 - beda 5 mix new.xlsx"
    try:
        containers = read_containers(input_file)
    except FileNotFoundError:
        print(f"File '{input_file}' tidak ditemukan!"); sys.exit(1)

    best, best_fit, history = optimize_ga(containers, pop_size=80, generations=50, verbose=True)

    cmap = {c.id: c for c in containers}
    vcg, lcg, tcg, n20_final, n40_final = cog_report(best, cmap)
    validate(best, cmap)

    base = os.path.splitext(os.path.basename(input_file))[0]
    export_excel(best, cmap, vcg, lcg, tcg, f"stowage_{base}_ga_only.xlsx")

    print(f"\n  Total waktu: {time.time()-t0:.1f} detik")

    print_unified_summary(
        "GA Only (tanpa SA)",
        n20_final, n40_final, vcg, tcg, lcg
    )